import yadisk
import openpyxl
from openpyxl import load_workbook

y = yadisk.YaDisk(token="y0_AgAAAABkLI6WAAhdVQAAAADQ0Vq6ANAYpuXzQ4qFGTe1gWFmyBCsmfo")
y.download('/123.xlsx', "123.xlsx")  # cкачивание файла
wb = openpyxl.load_workbook(filename="123.xlsx")
wb.active = 0
sheetVlad = wb.active
wb.close()

def skach():
    y.download('/123.xlsx', "123.xlsx")  # cкачивание файла

def obnovl():
    wb = openpyxl.load_workbook(filename="123.xlsx")
    wb.active = 0
    sheetVlad = wb.active


def nom_counter():
    a2 = (sheetVlad['A2'].value)   #0
    a3 = (sheetVlad['A3'].value)   #1
    a4 = (sheetVlad['A4'].value)   #2
    a5 = (sheetVlad['A5'].value)   #3
    a6 = (sheetVlad['A6'].value)   #4
    a7 = (sheetVlad['A7'].value)   #5
    a8 = (sheetVlad['A8'].value)   #6
    a9 = (sheetVlad['A9'].value)   #7
    a10 = (sheetVlad['A10'].value) #8
    a11 = (sheetVlad['A11'].value) #9
    a12 = (sheetVlad['A12'].value) #10
    a13 = (sheetVlad['A13'].value) #11
    a14 = (sheetVlad['A14'].value) #12
    a15 = (sheetVlad['A15'].value) #13
    a16 = (sheetVlad['A16'].value) #14
    a17 = (sheetVlad['A17'].value) #15
    a18 = (sheetVlad['A18'].value) #16
    a19 = (sheetVlad['A19'].value) #17
    a20 = (sheetVlad['A20'].value) #18
    a21 = (sheetVlad['A21'].value)  # 18
    a22 = (sheetVlad['A22'].value)  # 18
    a23 = (sheetVlad['A23'].value)  # 18
    a24 = (sheetVlad['A24'].value)  # 18
    a25 = (sheetVlad['A25'].value)  # 18
    a26 = (sheetVlad['A26'].value)  # 18
    a27 = (sheetVlad['A27'].value)  # 18
    a28 = (sheetVlad['A28'].value)  # 18
    a29 = (sheetVlad['A29'].value)  # 18
    a30 = (sheetVlad['A30'].value)  # 18
    a31 = (sheetVlad['A31'].value)  # 18
    a32 = (sheetVlad['A32'].value)  # 18
    a33 = (sheetVlad['A33'].value)  # 18
    a34 = (sheetVlad['A34'].value)  # 18
    a35 = (sheetVlad['A35'].value)  # 18
    a36 = (sheetVlad['A36'].value)  # 18
    a37 = (sheetVlad['A37'].value)  # 18
    a38 = (sheetVlad['A38'].value)  # 18
    a39 = (sheetVlad['A39'].value)  # 18
    a40 = (sheetVlad['A40'].value)  # 18
    a41 = (sheetVlad['A41'].value)  # 18
    a42 = (sheetVlad['A42'].value)  # 18
    a43 = (sheetVlad['A43'].value)  # 18
    a44 = (sheetVlad['A44'].value)  # 18
    a45 = (sheetVlad['A45'].value)  # 18
    a46 = (sheetVlad['A46'].value)  # 18
    a47 = (sheetVlad['A47'].value)  # 18
    a48 = (sheetVlad['A48'].value)  # 18
    a49 = (sheetVlad['A49'].value)  # 18
    a50 = (sheetVlad['A50'].value)  # 18
    a51 = (sheetVlad['A51'].value)  # 18
    a52 = (sheetVlad['A52'].value)  # 18
    a53 = (sheetVlad['A53'].value)  # 18
    a54 = (sheetVlad['A54'].value)  # 18
    a55 = (sheetVlad['A55'].value)  # 18
    a56 = (sheetVlad['A56'].value)  # 18
    a57 = (sheetVlad['A57'].value)  # 18
    a58 = (sheetVlad['A58'].value)  # 18
    a59 = (sheetVlad['A59'].value)  # 18
    a60 = (sheetVlad['A60'].value)  # 18
    a61 = (sheetVlad['A61'].value)  # 18
    a62 = (sheetVlad['A62'].value)  # 18
    a63 = (sheetVlad['A63'].value)  # 18
    a64 = (sheetVlad['A64'].value)  # 18
    a65 = (sheetVlad['A65'].value)  # 18
    a66 = (sheetVlad['A66'].value)  # 18
    a67 = (sheetVlad['A67'].value)  # 18
    a68 = (sheetVlad['A68'].value)  # 18
    a69 = (sheetVlad['A69'].value)  # 18
    a70 = (sheetVlad['A70'].value)  # 18
    a71 = (sheetVlad['A71'].value)  # 18
    a72 = (sheetVlad['A72'].value)  # 18
    a73 = (sheetVlad['A73'].value)  # 18
    a74 = (sheetVlad['A74'].value)  # 18
    a75 = (sheetVlad['A75'].value)  # 18
    a76 = (sheetVlad['A76'].value)  # 18
    a77 = (sheetVlad['A77'].value)  # 18
    a78 = (sheetVlad['A78'].value)  # 18
    a79 = (sheetVlad['A79'].value)  # 18
    a80 = (sheetVlad['A80'].value)  # 18
    a81 = (sheetVlad['A81'].value)  # 18
    a82 = (sheetVlad['A82'].value)  # 18
    a83 = (sheetVlad['A83'].value)  # 18
    a84 = (sheetVlad['A84'].value)  # 18
    a85 = (sheetVlad['A85'].value)  # 18
    a86 = (sheetVlad['A86'].value)  # 18
    a87 = (sheetVlad['A87'].value)  # 18
    a88 = (sheetVlad['A88'].value)  # 18
    a89 = (sheetVlad['A89'].value)  # 18
    a90 = (sheetVlad['A90'].value)  # 18
    a91 = (sheetVlad['A91'].value)  # 18
    a92 = (sheetVlad['A92'].value)  # 18
    a93 = (sheetVlad['A93'].value)  # 18
    a94 = (sheetVlad['A94'].value)  # 18
    a95 = (sheetVlad['A95'].value)  # 18
    a96 = (sheetVlad['A96'].value)  # 18
    a97 = (sheetVlad['A97'].value)  # 18
    a98 = (sheetVlad['A98'].value)  # 18
    a99 = (sheetVlad['A99'].value)  # 18
    a100 = (sheetVlad['A100'].value)  # 18
    a101 = (sheetVlad['A101'].value)
    a102 = (sheetVlad['A102'].value)
    a103 = (sheetVlad['A103'].value)
    a104 = (sheetVlad['A104'].value)
    a105 = (sheetVlad['A105'].value)
    a106 = (sheetVlad['A106'].value)
    a107 = (sheetVlad['A107'].value)
    a108 = (sheetVlad['A108'].value)
    a109 = (sheetVlad['A109'].value)
    a110 = (sheetVlad['A110'].value)
    a111 = (sheetVlad['A111'].value)
    a112 = (sheetVlad['A112'].value)
    a113 = (sheetVlad['A113'].value)
    a114 = (sheetVlad['A114'].value)
    a115 = (sheetVlad['A115'].value)
    a116 = (sheetVlad['A116'].value)
    a117 = (sheetVlad['A117'].value)
    a118 = (sheetVlad['A118'].value)
    a119 = (sheetVlad['A119'].value)
    a120 = (sheetVlad['A120'].value)
    a121 = (sheetVlad['A121'].value)
    a122 = (sheetVlad['A122'].value)
    a123 = (sheetVlad['A123'].value)
    a124 = (sheetVlad['A124'].value)
    a125 = (sheetVlad['A125'].value)
    a126 = (sheetVlad['A126'].value)
    a127 = (sheetVlad['A127'].value)
    a128 = (sheetVlad['A128'].value)
    a129 = (sheetVlad['A129'].value)
    a130 = (sheetVlad['A130'].value)
    a131 = (sheetVlad['A131'].value)
    a132 = (sheetVlad['A132'].value)
    a133 = (sheetVlad['A133'].value)
    a134 = (sheetVlad['A134'].value)
    a135 = (sheetVlad['A135'].value)
    a136 = (sheetVlad['A136'].value)
    a137 = (sheetVlad['A137'].value)
    a138 = (sheetVlad['A138'].value)
    a139 = (sheetVlad['A139'].value)
    a140 = (sheetVlad['A140'].value)
    a141 = (sheetVlad['A141'].value)
    a142 = (sheetVlad['A142'].value)
    a143 = (sheetVlad['A143'].value)
    a144 = (sheetVlad['A144'].value)
    a145 = (sheetVlad['A145'].value)
    a146 = (sheetVlad['A146'].value)
    a147 = (sheetVlad['A147'].value)
    a148 = (sheetVlad['A148'].value)
    a149 = (sheetVlad['A149'].value)
    a150 = (sheetVlad['A150'].value)

    return [a2, a3, a4, a5, a6, a7, a8, a9, a10, #9
            a11,a12,a13,a14,a15,a16,a17,a18,a19,a20, #10
            a21,a22,a23,a24,a25,a26,a27,a28,a29,a30,
            a31,a32,a33,a34,a35,a36,a37,a38,a39,a40,
            a41,a42,a43,a44,a45,a46,a47,a48,a49,a50,
            a51,a52,a53,a54,a55,a56,a57,a58,a59,a60,
            a61,a62,a63,a64,a65,a66,a67,a68,a69,a70,
            a71,a72,a73,a74,a75,a76,a77,a78,a79,a80,
            a81,a82,a83,a84,a85,a86,a87,a88,a89,a90,
            a91,a92,a93,a94,a95,a96,a97,a98,a99,a100, #90
            a101,a102,a103,a104,a105,a106,a107,a108,a109,a110, #50
            a111,a112,a113,a114,a115,a116,a117,a118,a119,a120,
            a121,a122,a123,a124,a125,a126,a127,a128,a129,a130,
            a131,a132,a133,a134,a135,a136,a137,a138,a139,a140,
            a141,a142,a143,a144,a145,a146,a147,a148,a149,a150]


def nom_name():
    wb = openpyxl.load_workbook(filename="123.xlsx")
    wb.active = 0
    sheetVlad = wb.active
    b2 = (sheetVlad['B2'].value)
    b3 = (sheetVlad['B3'].value)
    b4 = (sheetVlad['B4'].value)
    b5 = (sheetVlad['B5'].value)
    b6 = (sheetVlad['B6'].value)
    b7 = (sheetVlad['B7'].value)
    b8 = (sheetVlad['B8'].value)
    b9 = (sheetVlad['B9'].value)
    b10 = (sheetVlad['B10'].value)
    b11 = (sheetVlad['B11'].value)
    b12 = (sheetVlad['B12'].value)
    b13 = (sheetVlad['B13'].value)
    b14 = (sheetVlad['B14'].value)
    b15 = (sheetVlad['B15'].value)
    b16 = (sheetVlad['B16'].value)
    b17 = (sheetVlad['B17'].value)
    b18 = (sheetVlad['B18'].value)
    b19 = (sheetVlad['B19'].value)
    b20 = (sheetVlad['B20'].value)
    b21 = (sheetVlad['B21'].value)  # 18
    b22 = (sheetVlad['B22'].value)  # 18
    b23 = (sheetVlad['B23'].value)  # 18
    b24 = (sheetVlad['B24'].value)  # 18
    b25 = (sheetVlad['B25'].value)  # 18
    b26 = (sheetVlad['B26'].value)  # 18
    b27 = (sheetVlad['B27'].value)  # 18
    b28 = (sheetVlad['B28'].value)  # 18
    b29 = (sheetVlad['B29'].value)  # 18
    b30 = (sheetVlad['B30'].value)  # 18
    b31 = (sheetVlad['B31'].value)  # 18
    b32 = (sheetVlad['B32'].value)  # 18
    b33 = (sheetVlad['B33'].value)  # 18
    b34 = (sheetVlad['B34'].value)  # 18
    b35 = (sheetVlad['B35'].value)  # 18
    b36 = (sheetVlad['B36'].value)  # 18
    b37 = (sheetVlad['B37'].value)  # 18
    b38 = (sheetVlad['B38'].value)  # 18
    b39 = (sheetVlad['B39'].value)  # 18
    b40 = (sheetVlad['B40'].value)  # 18
    b41 = (sheetVlad['B41'].value)  # 18
    b42 = (sheetVlad['B42'].value)  # 18
    b43 = (sheetVlad['B43'].value)  # 18
    b44 = (sheetVlad['B44'].value)  # 18
    b45 = (sheetVlad['B45'].value)  # 18
    b46 = (sheetVlad['B46'].value)  # 18
    b47 = (sheetVlad['B47'].value)  # 18
    b48 = (sheetVlad['B48'].value)  # 18
    b49 = (sheetVlad['B49'].value)  # 18
    b50 = (sheetVlad['B50'].value)  # 18

    return [b2, b3, b4, b5, b6, b7, b8, b9, b10, b11, b12, b13, b14, b15, b16, b17, b18, b19, b20,b21,b22,b23,b24,b25,b26,b27,b28,b29,b30,b31,b32,b33,b34,b35,b36,b37,b38,b39,b40,b41,b42,b43,b44,b45,b46,b47,b48,b49,b50]







